import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# --- Step 1: Select input files ---
root = tk.Tk()
root.withdraw()

top_genes = filedialog.askopenfilename(
    title="Select Top Gene Lists Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
if not top_genes:
    raise SystemExit("No Top Genes file selected.")

annotations_file = filedialog.askopenfilename(
    title="Select eggNOG annotations file",
    filetypes=[("Excel files", "*.xlsx *.xls"), ("TSV files", "*.tsv"), ("CSV files", "*.csv")]
)
if not annotations_file:
    raise SystemExit("No annotations file selected.")

# --- Step 2: Load annotations (auto-detect header row for Excel, skiprows=2 for TSV/CSV) ---
if annotations_file.endswith(".tsv"):
    annotations = pd.read_csv(annotations_file, sep="\t", skiprows=2)
elif annotations_file.endswith(".csv"):
    annotations = pd.read_csv(annotations_file, skiprows=2)
else:  # Excel file
    raw = pd.read_excel(annotations_file, header=None)
    header_row = raw[0].eq("query").idxmax()
    annotations = pd.read_excel(annotations_file, header=header_row)

# --- Step 2a: Clean up mapping info ---
# Clean OGs
def clean_og(og_string):
    if pd.isna(og_string):
        return "No Assignment"
    first = og_string.split(",")[0]
    return first.split("@")[0]

annotations["Clean_OG"] = annotations["eggNOG_OGs"].apply(clean_og)

# Clean COG_category
if "COG_category" in annotations.columns:
    annotations["COG_category"] = annotations["COG_category"].replace("-", pd.NA)
    annotations["COG_category"] = annotations["COG_category"].fillna("No assignment")

# Dictionaries for mapping
og_map = dict(zip(annotations["query"], annotations["Clean_OG"]))
cog_map = annotations.set_index("query")["COG_category"].astype(str).to_dict()

# --- Step 3: Helper function to count COGs ---
def count_cogs(genes):
    cog_counts = {}
    for gene in genes.dropna():
        category = cog_map.get(gene, "No assignment")
        if category != "No assignment":
            for c in category:  # split multi-letter codes
                cog_counts[c] = cog_counts.get(c, 0) + 1
        else:
            cog_counts["No assignment"] = cog_counts.get("No assignment", 0) + 1
    return pd.Series(cog_counts)

# --- Step 4: Process Top50s and Top100s ---
xls = pd.ExcelFile(top_genes)
target_sheets = ["Top50s", "Top100s"]
book = load_workbook(top_genes)

summary_results = {}

with pd.ExcelWriter(top_genes, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    for sheet in target_sheets:
        if sheet not in xls.sheet_names:
            continue

        df = pd.read_excel(top_genes, sheet_name=sheet)

        # --- 4a. Count COGs for Summary ---
        for col in df.columns:
            summary_results[f"{sheet}_{col}"] = count_cogs(df[col])

        # --- 4b. Create OGs side-by-side ---
        og_df = df.map(lambda x: og_map.get(x, "No Assignment") if pd.notna(x) else x)

        combined = pd.DataFrame()
        for col in df.columns:
            combined[f"{col} (Loci)"] = df[col]
            combined[f"{col} (OGs)"] = og_df[col]

        og_sheet_name = f"{sheet}_OGs"
        if og_sheet_name in book.sheetnames:  # delete old sheet
            del book[og_sheet_name]

        combined.to_excel(writer, sheet_name=og_sheet_name, index=False)

    # --- Step 5: Write Summary ---
    summary_df = pd.DataFrame(summary_results).fillna(0).astype(int)

    if "No assignment" in summary_df.index:
        no_assign = summary_df.loc[["No assignment"]]
        summary_df = summary_df.drop(index="No assignment")
        summary_df = summary_df.sort_index()
        summary_df = pd.concat([summary_df, no_assign])

    summary_df.index.name = "COG_category"

    if "Summary" in book.sheetnames:  # delete old sheet
        del book["Summary"]

    summary_df.to_excel(writer, sheet_name="Summary")

print(f"✅ Summary, Top50s_OGs, and Top100s_OGs written to {top_genes}")
