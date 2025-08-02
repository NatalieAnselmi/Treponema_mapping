import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, numbers
from openpyxl.formula.translate import Translator

def clean_sample(s):
    if pd.isna(s):
        return ""
    s = str(s)
    s = s.replace(" ", "").upper()
    s = s.split("._")[0]
    return s

def map_health_status(status):
    if isinstance(status, str):
        if status.startswith("X"):
            return "Y"

        mappings = {
            "Healthy": "Healthy",
            "Diseased, Stable": "Stable PD",
            "Diseased; Stable": "Stable PD",
            "Diseased, Progressing": "Progressing PD",
            "Diseased; Progressing": "Progressing PD",
            "Diseased; Fluctuant": "Fluctuating PD",
            "Diseased; Fluctuating": "Fluctuating PD",
            "Diseased": "PD"
        }

        status = status.strip()
        for key in sorted(mappings.keys(), key=len, reverse=True):
            if status.startswith(key):
                return mappings[key]
        return status
    return status

def add_summary_rows(ws):
    data_start_row = 2
    data_end_row = ws.max_row
    header_row = 1
    count_row = data_end_row + 1
    percent_row = data_end_row + 2
    avg_row = data_end_row + 3
    sd_row = data_end_row + 4
    num_rows = data_end_row - 1  # Exclude header

    ws[f"B{count_row}"] = f"# >0 out of {num_rows}"
    ws[f"B{percent_row}"] = "Percent pos."
    ws[f"B{avg_row}"] = "Average"
    ws[f"B{sd_row}"] = "SD"

    bold_font = Font(bold=True)
    for row in [count_row, percent_row, avg_row, sd_row]:
        ws[f"B{row}"].font = bold_font

    first_data_col = 3  # Column C
    last_col = ws.max_column

    for col in range(first_data_col, last_col + 1):
        col_letter = get_column_letter(col)
        data_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"

        ws[f"{col_letter}{count_row}"] = f'=COUNTIF({data_range}, ">0")'
        percent_cell = ws[f"{col_letter}{percent_row}"]
        percent_cell.value = f"={col_letter}{count_row}/{num_rows}"
        percent_cell.number_format = '0.0%'  # Format as percent

        ws[f"{col_letter}{avg_row}"] = f'=AVERAGE({data_range})'
        ws[f"{col_letter}{sd_row}"] = f'=STDEV({data_range})'

def process_files(data_file_path, lookup_file_path, output_file_path):
    data_df = pd.read_excel(data_file_path)
    lookup_df = pd.read_excel(lookup_file_path)

    lookup_df['Sample_Cleaned'] = lookup_df['Sample'].apply(clean_sample)
    data_df['Sample_Cleaned'] = data_df['Sample'].apply(clean_sample)

    sample_to_health = dict(zip(
        lookup_df['Sample_Cleaned'],
        lookup_df['Health_Status'].apply(map_health_status)
    ))
    data_df['Mapped_Health_Status'] = data_df['Sample_Cleaned'].map(sample_to_health)
    data_df.drop(columns=['Sample_Cleaned'], inplace=True)

    sample_col_idx = data_df.columns.get_loc('Sample')
    cols = list(data_df.columns)
    cols.remove('Mapped_Health_Status')
    cols.insert(sample_col_idx + 1, 'Mapped_Health_Status')
    data_df = data_df[cols]

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        data_df.to_excel(writer, sheet_name='Full_Data', index=False)
        workbook = writer.book

        grouped = data_df.groupby('Mapped_Health_Status')
        for name, group in grouped:
            group.to_excel(writer, sheet_name=name[:31], index=False)

        unmatched = data_df[data_df['Mapped_Health_Status'].isna()]
        if not unmatched.empty:
            unmatched.to_excel(writer, sheet_name='Unmatched', index=False)

    # Add summary rows after save (openpyxl)
    wb = load_workbook(output_file_path)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Full_Data":
            continue
        ws = wb[sheet_name]
        add_summary_rows(ws)
    wb.save(output_file_path)

def launch_gui():
    root = tk.Tk()
    root.title("Health Status Mapper")
    root.geometry("500x300")

    def load_file(label):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            label.config(text=path)
            return path

    def process():
        if not lookup_label.cget("text") or not data_label.cget("text"):
            messagebox.showerror("Error", "Please select both files first.")
            return
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return
        try:
            process_files(data_label.cget("text"), lookup_label.cget("text"), output_path)
            messagebox.showinfo("Done", f"File saved to: {output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

    tk.Label(root, text="Select Sample Lookup File").pack(pady=5)
    tk.Button(root, text="Browse", command=lambda: load_file(lookup_label)).pack()
    lookup_label = tk.Label(root, text="")
    lookup_label.pack()

    tk.Label(root, text="Select Data File").pack(pady=10)
    tk.Button(root, text="Browse", command=lambda: load_file(data_label)).pack()
    data_label = tk.Label(root, text="")
    data_label.pack()

    tk.Button(root, text="Process Files", command=process, bg='green', fg='white').pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    launch_gui()
