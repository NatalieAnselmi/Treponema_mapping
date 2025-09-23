import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from tkinter import Tk, filedialog, messagebox, StringVar, BooleanVar
import tkinter as tk
from scipy.stats import f_oneway, kruskal, anderson
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import scikit_posthocs as sp
from Bio import SeqIO

# =========================
# Helpers for organizing code 
# =========================
def clean_sample(s):
    if pd.isna(s):
        return ""
    s = str(s)
    s = s.replace(" ", "").upper()
    s = s.split("._")[0]
    return s

def map_health_status(status):
    if isinstance(status, str):
        if status.startswith("X"):
            return "Y"
        mappings = {
            "Healthy": "Healthy",
            "Diseased, Stable": "Stable PD",
            "Diseased; Stable": "Stable PD",
            "Diseased, Progressing": "Progressing PD",
            "Diseased; Progressing": "Progressing PD",
            "Diseased; Fluctuant": "Fluctuating PD",
            "Diseased; Fluctuating": "Fluctuating PD",
            "Diseased": "PD"
        }
        status = status.strip()
        for key in sorted(mappings.keys(), key=len, reverse=True):
            if status.startswith(key):
                return mappings[key]
        return status
    return status

# =========================
# Pipeline: write base sheets (modified to accept locus->product mapping)
# =========================
def write_base_sheets(input_path, meta_path, output_path, locus_to_product=None):
    """Reads INPUT and METADATA, maps health status, writes:
       - Full_Data
       - One sheet per Mapped_Health_Status
       - Unmatched (if any)
       If locus_to_product is provided, replace column headers (from 3rd column onward)
       with the product names using that mapping before writing the sheets."""
    
    data_df = pd.read_excel(input_path)
    lookup_df = pd.read_excel(meta_path)

    # Prepare cleaned keys
    lookup_df['Sample_Cleaned'] = lookup_df['Sample'].apply(clean_sample)
    data_df['Sample_Cleaned'] = data_df['Sample'].apply(clean_sample)

    # Build mapping
    sample_to_health = dict(zip(
        lookup_df['Sample_Cleaned'],
        lookup_df['Health_Status'].apply(map_health_status)
    ))
    # Add mapped status
    data_df['Mapped_Health_Status'] = data_df['Sample_Cleaned'].map(sample_to_health)
    data_df.drop(columns=['Sample_Cleaned'], inplace=True)

    # Reorder column: put Mapped_Health_Status next to Sample
    if 'Sample' in data_df.columns and 'Mapped_Health_Status' in data_df.columns:
        sample_col_idx = data_df.columns.get_loc('Sample')
        cols = list(data_df.columns)
        cols.remove('Mapped_Health_Status')
        cols.insert(sample_col_idx + 1, 'Mapped_Health_Status')
        data_df = data_df[cols]

    # If mapping provided, create a copy with header replacement for writing
    def apply_header_mapping(df):
        if locus_to_product is None:
            return df
        df2 = df.copy()
        cols = list(df2.columns)

        # Track occurrences for numbering repeats
        name_counts = {}

        for i in range(2, len(cols)):  # 3rd col onward
            orig = cols[i]
            product = locus_to_product.get(str(orig), orig)

            # 1) Remove commas
            product = product.replace(",", "")

            # 2) Assign unique numbers for duplicates
            if product in name_counts:
                name_counts[product] += 1
                product = f"{product} {name_counts[product]}"
            else:
                name_counts[product] = 1

            cols[i] = product

        df2.columns = cols
        return df2

    # Write workbook (Full_Data + per-status + Unmatched) using mapped headers if requested
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Full_Data: write mapped header version
        apply_header_mapping(data_df).to_excel(writer, sheet_name='Full_Data', index=False)

        # Per-status sheets
        grouped = data_df.groupby('Mapped_Health_Status', dropna=False)
        for name, group in grouped:
            if pd.isna(name):
                continue
            safe_name = str(name)[:31] if name else "Unknown"
            apply_header_mapping(group).to_excel(writer, sheet_name=safe_name, index=False)

        # Unmatched
        unmatched = data_df[data_df['Mapped_Health_Status'].isna()]
        if not unmatched.empty:
            apply_header_mapping(unmatched).to_excel(writer, sheet_name='Unmatched', index=False)

    return output_path

# =========================
# Summary construction and Stats
# =========================
def get_gene_percent_bins():
    bins = np.array([0, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 100.1])
    labels = ["[0]", "(0-5]", "(5-10]", "(10-20]", "(20-30]", "(30-40]",
              "(40-50]", "(50-60]", "(60-70]", "(70-80]", "(80-90]", "(90-100]"]
    return bins, labels

def get_abundance_bins():
    bins = np.array([0, 1e-6, 0.1, 1, 5, 10, 20, 50, 100, 200, 500,
                     1000, 2000, 5000, 10000, 20000, 50000, 100000])
    labels = ["[0]", "(0-1e-6]", "(1e-6-0.1]", "(0.1-1]", "(1-5]",
              "(5-10]", "(10-20]", "(20-50]", "(50-100]", "(100-200]",
              "(200-500]", "(500-1000]", "(1000-2000]", "(2000-5000]",
              "(5000-10000]", "(10000-20000]", "(20000-50000]"]
    return bins, labels

def get_health_status_sheets(wb):
    return [ws.title for ws in wb.worksheets if ws.title not in ("Summary", "Full_Data", "Unmatched")]

def extract_status_gene_matrix(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    n_samples = max(0, max_row - 1)  # samples start at row 2
    n_genes = max(0, max_col - 2)    # genes start at col 3

    data = np.zeros((n_samples, n_genes), dtype=float)

    for col_idx in range(n_genes):
        col = col_idx + 3
        col_values = []
        for row in range(2, max_row + 1):
            val = ws.cell(row=row, column=col).value
            col_values.append(val if isinstance(val, (int, float)) else 0)
        if n_samples > 0:
            data[:, col_idx] = col_values

    return data

def extract_gene_names(ws):
    col_start = 3
    return [ws.cell(row=1, column=col).value for col in range(col_start, ws.max_column + 1)]

def compute_gene_stats_per_status(wb, health_statuses):
    gene_names = None
    percent_positive = {}
    mean_all = {}
    mean_nonzero = {}

    for status in health_statuses:
        ws = wb[status]
        if gene_names is None:
            gene_names = extract_gene_names(ws)
        data = extract_status_gene_matrix(ws)  # shape (samples, genes)

        if data.size == 0:
            pos_counts = np.zeros(len(gene_names), dtype=int)
            percent_pos = np.zeros(len(gene_names))
            mean_all_genes = np.zeros(len(gene_names))
            mean_nonzero_genes = np.zeros(len(gene_names))
        else:
            pos_counts = np.sum(data > 0, axis=0)
            percent_pos = (pos_counts / data.shape[0]) * 100 if data.shape[0] > 0 else 0
            mean_all_genes = np.mean(data, axis=0) if data.shape[0] > 0 else np.zeros(data.shape[1])

            with np.errstate(invalid='ignore', divide='ignore'):
                mean_nonzero_genes = np.true_divide(
                    np.sum(data, axis=0),
                    np.maximum(pos_counts, 1)
                )
                mean_nonzero_genes[pos_counts == 0] = 0

        percent_positive[status] = percent_pos
        mean_all[status] = mean_all_genes
        mean_nonzero[status] = mean_nonzero_genes

    return gene_names, percent_positive, mean_all, mean_nonzero

def write_summary_table(ws, gene_names, percent_positive, mean_all, mean_nonzero, health_statuses, start_col=1):
    ws.cell(row=1, column=start_col).value = "Gene"
    ws.cell(row=1, column=start_col).font = Font(bold=True)

    col = start_col + 1
    for status in health_statuses:
        ws.cell(row=1, column=col).value = status
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        ws.cell(row=1, column=col).font = Font(bold=True)

        ws.cell(row=2, column=col).value = "Percent pos."
        ws.cell(row=2, column=col + 1).value = "Mean (all)"
        ws.cell(row=2, column=col + 2).value = "Mean (non-zero)"
        for i in range(col, col + 3):
            ws.cell(row=2, column=i).font = Font(bold=True)
        col += 3

    row = 3
    for i, gene in enumerate(gene_names):
        ws.cell(row=row, column=start_col).value = gene
        ws.cell(row=row, column=start_col).font = Font(bold=True)

        for j, status in enumerate(health_statuses):
            status_col = start_col + 1 + j * 3
            ws.cell(row=row, column=status_col).value = round(percent_positive[status][i] / 100, 4)
            ws.cell(row=row, column=status_col).number_format = "0.00%"
            ws.cell(row=row, column=status_col + 1).value = round(float(mean_all[status][i]), 4)
            ws.cell(row=row, column=status_col + 2).value = round(float(mean_nonzero[status][i]), 4)
        row += 1

    return row  # first empty row after the table

def write_percent_positive_distribution(ws, percent_positive, health_statuses, bin_edges, bin_labels, start_row, start_col):
    ws.cell(row=start_row, column=start_col).value = "Percent positive samples"
    ws.cell(row=start_row, column=start_col).font = Font(bold=True)

    for i, status in enumerate(health_statuses):
        ws.cell(row=start_row, column=start_col + 1 + i).value = status
        ws.cell(row=start_row, column=start_col + 1 + i).font = Font(bold=True)

    data_list = [percent_positive[status] for status in health_statuses]
    data_matrix = np.column_stack(data_list) if len(data_list) > 0 else np.empty((0, 0))

    counts_matrix = np.zeros((len(bin_labels), len(health_statuses)), dtype=int)
    if data_matrix.size > 0:
        for col_idx in range(data_matrix.shape[1]):
            counts, _ = np.histogram(data_matrix[:, col_idx], bins=bin_edges)
            counts_matrix[:, col_idx] = counts

    for idx, label in enumerate(bin_labels):
        row = start_row + 1 + idx
        ws.cell(row=row, column=start_col).value = label
        ws.cell(row=row, column=start_col).font = Font(bold=True)
        for c, count in enumerate(counts_matrix[idx], start=start_col + 1):
            ws.cell(row=row, column=c).value = int(count)

    return start_row + len(bin_labels) + 1

def write_abundance_distribution(ws, wb, health_statuses, bins, labels, start_row, start_col):
    ws.cell(row=start_row, column=start_col).value = "Abundance distribution"
    ws.cell(row=start_row, column=start_col).font = Font(bold=True)

    for i, status in enumerate(health_statuses):
        ws.cell(row=start_row, column=start_col + 1 + i).value = status
        ws.cell(row=start_row, column=start_col + 1 + i).font = Font(bold=True)

    data_list = []
    for status in health_statuses:
        ws_status = wb[status]
        max_row = ws_status.max_row - 4 if ws_status.max_row >= 6 else ws_status.max_row
        max_col = ws_status.max_column

        vals = []
        for row in range(2, max_row + 1):
            for col in range(3, max_col + 1):
                val = ws_status.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    vals.append(val)
        data_list.append(np.array(vals, dtype=float))

    counts_matrix = np.vstack([
        np.histogram(data_list[i], bins=bins)[0] if data_list[i].size > 0 else np.zeros(len(labels), dtype=int)
        for i in range(len(health_statuses))
    ]).T

    for i, label in enumerate(labels):
        row = start_row + 1 + i
        ws.cell(row=row, column=start_col).value = label
        ws.cell(row=row, column=start_col).font = Font(bold=True)
        for c, count in enumerate(counts_matrix[i], start=start_col + 1):
            ws.cell(row=row, column=c).value = int(count)

    return start_row + len(labels) + 1

def p_to_stars(p):
    if p > 0.05:
        return "ns"
    elif p <= 0.0001:
        return "****"
    elif p <= 0.001:
        return "***"
    elif p <= 0.01:
        return "**"
    elif p <= 0.05:
        return "*"

def write_posthoc_results(ws, posthoc_result, test_type, start_row, start_col):
    ws.cell(row=start_row, column=start_col).value = "Post-hoc Test Results"
    ws.cell(row=start_row, column=start_col).font = Font(bold=True)

    if test_type.startswith("ANOVA") and posthoc_result is not None:
        rows = posthoc_result.summary().data
        header = rows[0]
        data_rows = rows[1:]
        for c, heading in enumerate(header):
            ws.cell(row=start_row + 1, column=start_col + c).value = heading
            ws.cell(row=start_row + 1, column=start_col + c).font = Font(bold=True)
        for r, row_data in enumerate(data_rows):
            for c, val in enumerate(row_data):
                ws.cell(row=start_row + 2 + r, column=start_col + c).value = val
        return start_row + 2 + len(data_rows)

    elif test_type == "Kruskal-Wallis" and posthoc_result is not None:
        cols = list(posthoc_result.columns)
        ws.cell(row=start_row + 1, column=start_col).value = ""
        for c, colname in enumerate(cols):
            ws.cell(row=start_row + 1, column=start_col + 1 + c).value = colname
            ws.cell(row=start_row + 1, column=start_col + 1 + c).font = Font(bold=True)

        for r, index in enumerate(posthoc_result.index):
            ws.cell(row=start_row + 2 + r, column=start_col).value = index
            ws.cell(row=start_row + 2 + r, column=start_col).font = Font(bold=True)
            for c, colname in enumerate(cols):
                p_val = float(posthoc_result.at[index, colname])
                star = p_to_stars(p_val)
                formatted = f"{star} ({p_val:.4f})"
                ws.cell(row=start_row + 2 + r, column=start_col + 1 + c).value = formatted

        return start_row + 2 + len(posthoc_result)

    else:
        ws.cell(row=start_row + 1, column=start_col).value = "No post-hoc test performed."
        return start_row + 2

def test_normality_and_anova(status_values):
    normality = {}
    lognormality = {}

    for status, arr in status_values.items():
        arr = np.asarray(arr, dtype=float)
        if arr.size == 0:
            normality[status] = False
            lognormality[status] = False
            continue

        # Anderson–Darling on raw
        try:
            result = anderson(arr, dist='norm')
            crit_val_5pct = result.critical_values[2]
            normality[status] = (result.statistic < crit_val_5pct)
        except Exception:
            normality[status] = False

        # Anderson–Darling on log (>0)
        log_arr = np.log(arr[arr > 0])
        if log_arr.size == 0:
            lognormality[status] = False
        else:
            try:
                result_log = anderson(log_arr, dist='norm')
                crit_val_5pct_log = result_log.critical_values[2]
                lognormality[status] = (result_log.statistic < crit_val_5pct_log)
            except Exception:
                lognormality[status] = False

    all_normal = all(normality.values()) if normality else False
    all_lognormal = all(lognormality.values()) if lognormality else False

    if all_normal:
        anova_type = 'ANOVA'
        anova_result = f_oneway(*[np.asarray(status_values[st], dtype=float) for st in status_values])
    elif all_lognormal:
        anova_type = 'ANOVA (log-transformed)'
        anova_result = f_oneway(*[np.log(np.asarray(status_values[st], dtype=float)) for st in status_values])
    else:
        anova_type = 'Kruskal-Wallis'
        anova_result = kruskal(*[np.asarray(status_values[st], dtype=float) for st in status_values])

    posthoc = None
    if hasattr(anova_result, "pvalue") and anova_result.pvalue < 0.05:
        combined = []
        groups = []
        for status in status_values:
            arr = np.asarray(status_values[status], dtype=float)
            if anova_type == 'ANOVA (log-transformed)':
                arr = np.log(arr[arr > 0])
            combined.extend(arr)
            groups.extend([status] * len(arr))
        combined = np.array(combined, dtype=float)
        groups = np.array(groups, dtype=object)

        if anova_type.startswith('ANOVA'):
            try:
                posthoc = pairwise_tukeyhsd(combined, groups)
            except Exception:
                posthoc = None
        else:
            try:
                df = pd.DataFrame({'group': groups, 'value': combined})
                posthoc = sp.posthoc_dunn(df, val_col='value', group_col='group', p_adjust='fdr_bh')
            except Exception:
                posthoc = None

    return {
        'normality': normality,
        'lognormality': lognormality,
        'anova_type': anova_type,
        'anova_result': anova_result,
        'posthoc': posthoc
    }

def write_abundance_stats(ws, wb, health_statuses, start_row, start_col):
    status_values = {status: [] for status in health_statuses}

    for ws_status in wb.worksheets:
        if ws_status.title not in health_statuses:
            continue
        max_row = ws_status.max_row - 4 if ws_status.max_row >= 6 else ws_status.max_row
        max_col = ws_status.max_column
        for row in range(2, max_row + 1):
            for col in range(3, max_col + 1):
                val = ws_status.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    status_values[ws_status.title].append(float(val))

    for status in status_values:
        status_values[status] = np.array(status_values[status], dtype=float)

    stat_blocks = [
        ("Abundance Stats (All)", lambda arr: arr),
        ("Abundance Stats (Non-zero)", lambda arr: arr[arr > 0])
    ]
    stats_labels = ["min", "1st quart", "median", "3rd quart", "max"]

    current_row = start_row
    for block_title, filter_func in stat_blocks:
        ws.cell(row=current_row, column=start_col).value = block_title
        ws.cell(row=current_row, column=start_col).font = Font(bold=True)

        for j, status in enumerate(health_statuses):
            ws.cell(row=current_row, column=start_col + 1 + j).value = status
            ws.cell(row=current_row, column=start_col + 1 + j).font = Font(bold=True)

        for i, stat_label in enumerate(stats_labels):
            row = current_row + 1 + i
            ws.cell(row=row, column=start_col).value = stat_label
            ws.cell(row=row, column=start_col).font = Font(bold=True)

            for j, status in enumerate(health_statuses):
                arr = filter_func(status_values[status])
                if arr.size == 0:
                    value = ""
                elif stat_label == "min":
                    value = np.min(arr)
                elif stat_label == "1st quart":
                    value = np.percentile(arr, 25)
                elif stat_label == "median":
                    value = np.median(arr)
                elif stat_label == "3rd quart":
                    value = np.percentile(arr, 75)
                elif stat_label == "max":
                    value = np.max(arr)
                ws.cell(row=row, column=start_col + 1 + j).value = (
                    round(float(value), 3) if value != "" else ""
                )
        current_row += len(stats_labels) + 2

    nonzero_status_values = {
        status: status_values[status][status_values[status] > 0]
        for status in health_statuses
    }

    test_results = test_normality_and_anova(nonzero_status_values)

    summary_start = current_row + 1
    ws.cell(row=summary_start, column=start_col).value = "Normality Test (Anderson-Darling)"
    ws.cell(row=summary_start, column=start_col).font = Font(bold=True)
    row = summary_start + 1
    for status in health_statuses:
        norm = "Yes" if test_results['normality'].get(status, False) else "No"
        ws.cell(row=row, column=start_col).value = f"{status}: {norm}"
        row += 1

    row += 1
    ws.cell(row=row, column=start_col).value = "Lognormality Test (Anderson-Darling on log data)"
    ws.cell(row=row, column=start_row).font = Font(bold=True)
    row += 1
    for status in health_statuses:
        lognorm = "Yes" if test_results['lognormality'].get(status, False) else "No"
        ws.cell(row=row, column=start_col).value = f"{status}: {lognorm}"
        row += 1

    row += 1
    ws.cell(row=row, column=start_col).value = f"Chosen test: {test_results['anova_type']}"
    ws.cell(row=row, column=start_col).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=start_col).value = f"Test statistic: {round(float(test_results['anova_result'].statistic), 4)}"
    row += 1
    ws.cell(row=row, column=start_col).value = f"p-value: {round(float(test_results['anova_result'].pvalue), 6)}"

    posthoc_end_row = write_posthoc_results(
        ws, test_results['posthoc'], test_results['anova_type'], row + 2, start_col
    )
    return posthoc_end_row

# =========================
# Pos. Genes 
# =========================
def get_positive_genes(ws, health_statuses):
    header_row = 1
    subheader_row = 2
    gene_col = 1  # Column A holds gene names

    positive_genes = {status: [] for status in health_statuses}

    for col in range(1, ws.max_column + 1):
        cell_value = str(ws.cell(row=header_row, column=col).value).strip() if ws.cell(row=header_row, column=col).value else None
        if cell_value in health_statuses:
            if str(ws.cell(row=subheader_row, column=col).value).strip() == "Percent pos.":
                status = cell_value
                for row in range(subheader_row + 1, ws.max_row + 1):
                    p_val = ws.cell(row=row, column=col).value
                    if p_val is not None and isinstance(p_val, (int, float)) and p_val > 0:
                        gene_name = ws.cell(row=row, column=gene_col).value
                        if gene_name:
                            positive_genes[status].append(gene_name)
    return positive_genes

def write_positive_genes(wb, positive_genes):
    if "Pos. Genes" in wb.sheetnames:
        del wb["Pos. Genes"]
    ws_new = wb.create_sheet("Pos. Genes")

    for col_idx, status in enumerate(positive_genes.keys(), start=1):
        ws_new.cell(row=1, column=col_idx).value = status
        ws_new.cell(row=1, column=col_idx).font = Font(bold=True)
        for row_idx, gene in enumerate(positive_genes[status], start=2):
            ws_new.cell(row=row_idx, column=col_idx).value = gene

# =========================
# GBFF parsing helper
# =========================
def build_locus_to_product_map(gbff_path):
    locus_to_product = {}
    for record in SeqIO.parse(gbff_path, "genbank"):
        for feature in record.features:
            if feature.type == "CDS":
                locus_tags = feature.qualifiers.get("locus_tag", [])
                products = feature.qualifiers.get("product", [])
                if locus_tags and products:
                    locus_to_product[locus_tags[0]] = products[0]
    return locus_to_product

# =========================
# Main 
# =========================
def main():
    # ---- GUI window ----
    root = tk.Tk()
    root.title("Run Pipeline")
    root.geometry("520x320")

    input_path_var = StringVar(value="")
    meta_path_var = StringVar(value="")
    gbff_path_var = StringVar(value="")
    convert_var = BooleanVar(value=False)

    def choose_input():
        p = filedialog.askopenfilename(title="Select INPUT Excel (RPKM matrix)",
                                       filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            input_path_var.set(p)
            btn_input.config(text=os.path.basename(p))
            check_enable_run()

    def choose_meta():
        p = filedialog.askopenfilename(title="Select METADATA Excel",
                                       filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            meta_path_var.set(p)
            btn_meta.config(text=os.path.basename(p))
            check_enable_run()

    def choose_gbff():
        p = filedialog.askopenfilename(title="Select .gbff (GenBank) file",
                                       filetypes=[("GBFF files", "*.gbff *.gbk"), ("All files", "*.*")])
        if p:
            gbff_path_var.set(p)
            gbff_button.config(text=os.path.basename(p))
            check_enable_run()

    def toggle_gbff():
        if convert_var.get():
            gbff_button.pack(anchor="w", pady=5)
        else:
            gbff_button.pack_forget()
            gbff_path_var.set("")
            gbff_button.config(text="Select GBFF")
        check_enable_run()

    def check_enable_run():
        inp = bool(input_path_var.get())
        meta = bool(meta_path_var.get())
        if convert_var.get():
            gb = bool(gbff_path_var.get())
            enabled = inp and meta and gb
        else:
            enabled = inp and meta
        btn_run.config(state="normal" if enabled else "disabled")

    # Title label
    tk.Label(root, text="Select required files and options", font=("Arial", 12, "bold")).pack(pady=8)

    # Frame for Input/Metadata buttons
    frame_files = tk.Frame(root)
    frame_files.pack(pady=6)

    btn_input = tk.Button(frame_files, text="Select Input", width=30, command=choose_input)
    btn_input.grid(row=0, column=0, padx=8, pady=6)

    btn_meta = tk.Button(frame_files, text="Select Metadata", width=30, command=choose_meta)
    btn_meta.grid(row=1, column=0, padx=8, pady=6)

    # Frame for checkbox and GBFF button
    checkbox_frame = tk.Frame(root)
    checkbox_frame.pack(pady=10)

    convert_checkbox = tk.Checkbutton(checkbox_frame, text="Convert gene loci to gene products?",
                                      variable=convert_var, command=toggle_gbff)
    convert_checkbox.pack(anchor="w")

    gbff_button = tk.Button(checkbox_frame, text="Select GBFF", command=choose_gbff)
    gbff_button.pack_forget()  # Hidden until checkbox toggled

    # Run button at bottom
    btn_run = tk.Button(root, text="Run pipeline", bg="lightgreen", state="disabled")
    btn_run.pack(pady=10)

    def run_pipeline():
        input_path = input_path_var.get()
        meta_path = meta_path_var.get()
        gbff_path = gbff_path_var.get() if convert_var.get() else None

        if not input_path:
            messagebox.showerror("Missing file", "No INPUT file selected.")
            return
        if not meta_path:
            messagebox.showerror("Missing file", "No METADATA file selected.")
            return
        if convert_var.get() and not gbff_path:
            messagebox.showerror("Missing file", "Conversion selected but no GBFF provided.")
            return

        default_out = os.path.splitext(os.path.basename(input_path))[0] + "_OUTPUT.xlsx"
        out_path = filedialog.asksaveasfilename(title="Save output workbook as...",
                                                defaultextension=".xlsx",
                                                initialfile=default_out,
                                                filetypes=[("Excel files", "*.xlsx")])
        if not out_path:
            return

        # Build locus->product map if requested
        locus_map = None
        if convert_var.get():
            try:
                locus_map = build_locus_to_product_map(gbff_path)
                if not locus_map:
                    messagebox.showwarning("Warning", "No locus->product mappings found in GBFF. Will keep loci as-is.")
                    locus_map = None
            except Exception as e:
                messagebox.showerror("Error parsing GBFF", f"Error while parsing GBFF:\n{e}")
                return

        # Run existing pipeline steps
        write_base_sheets(input_path, meta_path, out_path, locus_to_product=locus_map)
        messagebox.showinfo("Done", "Base sheets written.")

        # Rest of pipeline: open workbook and build Summary & Pos. Genes
        wb = openpyxl.load_workbook(out_path, data_only=True)

        # Remove Summary if exists and rebuild it
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        summary_ws = wb.create_sheet("Summary")

        # Determine available health status sheets
        health_status_sheets = get_health_status_sheets(wb)
        preferred_order = ["Healthy", "PD", "Stable PD", "Fluctuating PD", "Progressing PD"]
        ordered = [s for s in preferred_order if s in health_status_sheets]
        others = [s for s in health_status_sheets if s not in ordered]
        health_status_sheets = ordered + sorted(others)

        # Build Summary
        gene_names, percent_positive, mean_all, mean_nonzero = compute_gene_stats_per_status(wb, health_status_sheets)
        percent_bins, percent_labels = get_gene_percent_bins()
        abundance_bins, abundance_labels = get_abundance_bins()

        # Main gene table at left
        last_summary_row = write_summary_table(
            summary_ws, gene_names, percent_positive, mean_all, mean_nonzero, health_status_sheets, start_col=1
        )

        # Percent positive distribution (to the right of the gene table)
        summary_col_width = 1 + 3 * len(health_status_sheets)
        abundance_start_col = summary_col_width + 2

        percent_end_row = write_percent_positive_distribution(
            summary_ws, percent_positive, health_status_sheets,
            percent_bins, percent_labels,
            start_row=1,
            start_col=abundance_start_col
        )

        # Abundance distribution (below percent positive distribution)
        abundance_start_row = percent_end_row + 2
        abundance_end_row = write_abundance_distribution(
            summary_ws, wb, health_status_sheets,
            abundance_bins, abundance_labels,
            start_row=abundance_start_row,
            start_col=abundance_start_col
        )

        # Abundance stats + AD normality/lognormality + chosen test + posthoc
        stats_start_row = abundance_end_row + 2
        _ = write_abundance_stats(
            summary_ws, wb, health_status_sheets,
            start_row=stats_start_row,
            start_col=abundance_start_col
        )

        # Pos. Genes
        positive = get_positive_genes(summary_ws, health_status_sheets)
        write_positive_genes(wb, positive)

        # Save workbook
        wb.save(out_path)
        messagebox.showinfo("Finished", f"Output written to:\n{out_path}")
        root.quit()

    btn_run.config(command=run_pipeline)
    btn_run.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
