import pandas as pd
import numpy as np
import os, re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

PSEUDOCOUNT = 0.1
SHEETS = ["Healthy", "PD", "Stable PD", "Fluctuating PD", "Progressing PD"]

def normalize_key(name: str) -> str:
    if pd.isna(name):
        return ""
    name = re.sub(r"[^\w]", "_", str(name))
    name = re.sub(r"_+", "_", name).strip("_")
    return name

def clean_og(og_string):
    if pd.isna(og_string):
        return "No assignment"
    return og_string.split(",")[0].split("@")[0]

def load_annotations(mapper_path):
    if mapper_path.endswith(".tsv"):
        annotations = pd.read_csv(mapper_path, sep="\t", comment="#")
    elif mapper_path.endswith(".csv"):
        annotations = pd.read_csv(mapper_path, comment="#")
    else:
        raw = pd.read_excel(mapper_path, header=None)
        header_row = raw[0].eq("query").idxmax()
        annotations = pd.read_excel(mapper_path, header=header_row)
    annotations["Clean_OG"] = annotations["eggNOG_OGs"].apply(clean_og)
    annotations["COG_category"] = annotations["COG_category"].replace("-", pd.NA).fillna("No assignment")
    annotations["norm_query"] = annotations["query"].apply(normalize_key)
    cog_map = annotations.set_index("norm_query")["COG_category"].astype(str).to_dict()
    return cog_map

def count_cogs_for_sheet(df, cog_map):
    """Count COGs for every column in a dataframe and return summary table."""
    results = {}
    for col in df.columns:
        counts = {}
        for g in df[col].dropna():
            key = normalize_key(g)
            cat = cog_map.get(key, "No assignment")
            if cat != "No assignment":
                for c in str(cat):
                    counts[c] = counts.get(c, 0) + 1
            else:
                counts["No assignment"] = counts.get("No assignment", 0) + 1
        results[col] = counts
    summary = pd.DataFrame(results).fillna(0).astype(int)
    # Force "No assignment" to bottom
    if "No assignment" in summary.index:
        no_assign = summary.loc[["No assignment"]]
        summary = summary.drop("No assignment", errors="ignore").sort_index()
        summary = pd.concat([summary, no_assign])
    summary.index.name = "COG_category"
    return summary

def analyze_file(file_path, mapper_path):
    dfs = {s: pd.read_excel(file_path, sheet_name=s) for s in SHEETS}
    dfs = {s: df.rename(columns={c: normalize_key(c) for c in df.columns}) for s, df in dfs.items()}

    top50_abundant, top100_abundant = {}, {}
    for s, df in dfs.items():
        abundance = df.iloc[:, 2:].sum(axis=0).sort_values(ascending=False)
        top50_abundant[s] = abundance.head(50).index.tolist()
        top100_abundant[s] = abundance.head(100).index.tolist()

    combined = pd.concat([df.iloc[:, 2:] for df in dfs.values()], axis=0)
    variability = combined.var(axis=0).sort_values(ascending=False)
    top50_var = variability.head(50).index.tolist()
    top100_var = variability.head(100).index.tolist()

    means = {s: df.iloc[:, 2:].mean(axis=0) for s, df in dfs.items()}
    comps = ["PD", "Stable PD", "Fluctuating PD", "Progressing PD"]
    top50_vh, top100_vh = {}, {}
    for c in comps:
        log2fc = np.log2((means[c] + PSEUDOCOUNT) / (means["Healthy"] + PSEUDOCOUNT))
        top50_vh[f"{normalize_key(c)}_Higher"] = log2fc.nlargest(50).index
        top50_vh[f"{normalize_key(c)}_Lower"] = log2fc.nsmallest(50).index
        top100_vh[f"{normalize_key(c)}_Higher"] = log2fc.nlargest(100).index
        top100_vh[f"{normalize_key(c)}_Lower"] = log2fc.nsmallest(100).index

    out_path = file_path.replace(".xlsx", "_Tops.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({
            "Healthy": pd.Series(top50_abundant["Healthy"]),
            "PD": pd.Series(top50_abundant["PD"]),
            "Stable PD": pd.Series(top50_abundant["Stable PD"]),
            "Fluctuating PD": pd.Series(top50_abundant["Fluctuating PD"]),
            "Progressing PD": pd.Series(top50_abundant["Progressing PD"]),
            "Most_Variable": pd.Series(top50_var),
        }).to_excel(writer, sheet_name="Top50s", index=False)

        pd.DataFrame({
            "Healthy": pd.Series(top100_abundant["Healthy"]),
            "PD": pd.Series(top100_abundant["PD"]),
            "Stable PD": pd.Series(top100_abundant["Stable PD"]),
            "Fluctuating PD": pd.Series(top100_abundant["Fluctuating PD"]),
            "Progressing PD": pd.Series(top100_abundant["Progressing PD"]),
            "Most_Variable": pd.Series(top100_var),
        }).to_excel(writer, sheet_name="Top100s", index=False)

        pd.DataFrame({k: pd.Series(v) for k, v in top50_vh.items()}).to_excel(
            writer, sheet_name="Top50vHealth", index=False)
        pd.DataFrame({k: pd.Series(v) for k, v in top100_vh.items()}).to_excel(
            writer, sheet_name="Top100vHealth", index=False)

    # ---- Build COG Summary ----
    cog_map = load_annotations(mapper_path)
    xls = pd.ExcelFile(out_path)

    summary_tables = {}
    # Normal Top50/Top100
    for sheet in ["Top50s", "Top100s"]:
        df = pd.read_excel(out_path, sheet_name=sheet)
        df = df.map(normalize_key)
        summary_tables[sheet] = count_cogs_for_sheet(df, cog_map)

    # vHealth sheets split
    for sheet in ["Top50vHealth", "Top100vHealth"]:
        df = pd.read_excel(out_path, sheet_name=sheet)
        df = df.map(normalize_key)
        higher_cols = [c for c in df.columns if c.endswith("_Higher")]
        lower_cols = [c for c in df.columns if c.endswith("_Lower")]
        if higher_cols:
            summary_tables[f"{sheet}_Higher"] = count_cogs_for_sheet(df[higher_cols], cog_map)
        if lower_cols:
            summary_tables[f"{sheet}_Lower"] = count_cogs_for_sheet(df[lower_cols], cog_map)

    wb = load_workbook(out_path)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    # Write tables in 2x3 grid dynamically
    grid = [
        ["Top50s", "Top50vHealth_Higher", "Top100vHealth_Higher"],
        ["Top100s", "Top50vHealth_Lower", "Top100vHealth_Lower"],
    ]

    row_cursor = 1
    for row_tables in grid:
        col_cursor = 1
        row_block_height = 0
        for tbl_name in row_tables:
            if tbl_name not in summary_tables:
                continue
            df = summary_tables[tbl_name].reset_index()
            # Write title
            ws.cell(row=row_cursor, column=col_cursor, value=tbl_name)
            # Write dataframe
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, val in enumerate(row, start=0):
                    ws.cell(row=row_cursor + r_idx, column=col_cursor + c_idx, value=val)
            # Track max table height
            row_block_height = max(row_block_height, len(df) + 2)
            # Move col cursor right (width of table + 2 blank cols)
            col_cursor += len(df.columns) + 2
        # After finishing this row of tables, move row cursor down
        row_cursor += row_block_height + 1

    wb.save(out_path)
    print(f"✅ Saved: {out_path}")

# ---------- GUI ----------
root = tk.Tk()
root.title("Gene + COG Analysis")
root.geometry("550x600")

file_buttons, mapper_buttons, selected_paths, selected_mappers = [], [], [], []

def check_all_selected():
    if all(selected_paths) and all(selected_mappers):
        run_button.pack(pady=20)
    else:
        run_button.forget()

def browse_file(idx):
    p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if p:
        selected_paths[idx] = p
        file_buttons[idx].config(text=os.path.basename(p))
        check_all_selected()

def browse_mapper(idx):
    m = filedialog.askopenfilename(filetypes=[("Excel/TSV/CSV", "*.xlsx *.xls *.tsv *.csv")])
    if m:
        selected_mappers[idx] = m
        mapper_buttons[idx].config(text=os.path.basename(m))
        check_all_selected()

def load_file_inputs(count):
    for w in file_input_frame.winfo_children():
        w.destroy()
    file_buttons.clear(); mapper_buttons.clear(); selected_paths.clear(); selected_mappers.clear()
    for i in range(count):
        f = tk.Frame(file_input_frame); f.pack(pady=5)
        tk.Label(f, text=f"Input {i+1}:").pack(side="left", padx=5)
        b1 = tk.Button(f, text="Browse Excel", command=lambda idx=i: browse_file(idx))
        b1.pack(side="left", padx=5)
        b2 = tk.Button(f, text="Browse Mapper", command=lambda idx=i: browse_mapper(idx))
        b2.pack(side="left", padx=5)
        file_buttons.append(b1); mapper_buttons.append(b2)
        selected_paths.append(None); selected_mappers.append(None)
    check_all_selected()

def confirm_count():
    try:
        c = int(file_count_entry.get())
        if c < 1: raise ValueError
        load_file_inputs(c)
    except ValueError:
        messagebox.showerror("Invalid input", "Please enter an integer ≥ 1.")

def run_all():
    for p, m in zip(selected_paths, selected_mappers):
        if p and m:
            analyze_file(p, m)
    messagebox.showinfo("Done", "All analyses complete.")

tk.Label(root, text="How many files to analyze?").pack(pady=10)
file_count_entry = tk.Entry(root, justify="center"); file_count_entry.pack()
tk.Button(root, text="Confirm", command=confirm_count).pack(pady=5)
file_input_frame = tk.Frame(root); file_input_frame.pack(pady=10)
run_button = tk.Button(root, text="Run Analysis", fg="blue", command=run_all)

root.mainloop()
