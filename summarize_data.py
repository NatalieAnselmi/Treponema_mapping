import openpyxl
from openpyxl.styles import Font, numbers
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog

def bin_gene_percents(values):
    bins = [0, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    labels = ["[0]", "(0, 5]", "(5, 10]", "(10, 20]", "(20, 30]", "(30, 40]",
              "(40, 50]", "(50, 60]", "(60, 70]", "(70, 80]", "(80, 90]", "(90, 100)", "[100]"]
    bin_edges = [(-0.01, 0), (0, 5), (5, 10), (10, 20), (20, 30), (30, 40), (40, 50),
                 (50, 60), (60, 70), (70, 80), (80, 90), (90, 100), (100, 100.01)]

    bin_counts = {label: 0 for label in labels}

    for val in values:
        val = val * 100 if 0 < val <= 1 else val  # Handle 0.25 vs 25
        for (low, high), label in zip(bin_edges, labels):
            if low < val <= high:
                bin_counts[label] += 1
                break

    return bin_counts

def select_excel_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File", filetypes=[("Excel files", "*.xlsx")]
    )
    return file_path

def extract_gene_summary_data(wb):
    gene_data = {}  # {gene: {health_status: [count, percent, avg, sd]}}
    count_labels = {}
    health_statuses = []

    for ws in wb.worksheets:
        if ws.title in ("Summary", "Full_Data"):
            continue
        status = ws.title
        health_statuses.append(status)

        col_start = 3
        num_cols = ws.max_column
        gene_names = [ws.cell(row=1, column=col).value for col in range(col_start, num_cols + 1)]

        row_count = ws.max_row
        counts_row = row_count - 3
        percent_row = row_count - 2
        avg_row = row_count - 1
        sd_row = row_count

        count_labels[status] = ws.cell(row=counts_row, column=2).value

        for i, gene in enumerate(gene_names):
            if gene not in gene_data:
                gene_data[gene] = {}

            col = col_start + i
            count = ws.cell(row=counts_row, column=col).value
            percent = ws.cell(row=percent_row, column=col).value
            avg = ws.cell(row=avg_row, column=col).value
            sd = ws.cell(row=sd_row, column=col).value

            count = count if isinstance(count, (int, float)) else 0
            percent = round(percent, 4) if isinstance(percent, (int, float)) else 0
            avg = round(avg, 3) if isinstance(avg, (int, float)) else 0
            sd = round(sd, 3) if isinstance(sd, (int, float)) else 0

            gene_data[gene][status] = [count, percent, avg, sd]

    return gene_data, count_labels, health_statuses

def write_summary_table(ws, gene_data, count_labels, health_statuses, start_col=1):
    ws.cell(row=1, column=start_col).value = "Gene"
    ws.cell(row=1, column=start_col).font = Font(bold=True)

    # Header setup
    col = start_col + 1
    for status in health_statuses:
        ws.cell(row=1, column=col).value = status
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        ws.cell(row=1, column=col).font = Font(bold=True)

        ws.cell(row=2, column=col).value = count_labels[status]
        ws.cell(row=2, column=col + 1).value = "Percent pos."
        ws.cell(row=2, column=col + 2).value = "Average"
        ws.cell(row=2, column=col + 3).value = "SD"

        for i in range(col, col + 4):
            ws.cell(row=2, column=i).font = Font(bold=True)
        col += 4

    # Data rows
    row = 3
    for gene in sorted(gene_data.keys()):
        ws.cell(row=row, column=start_col).value = gene
        ws.cell(row=row, column=start_col).font = Font(bold=True)

        for i, status in enumerate(health_statuses):
            values = gene_data[gene].get(status, [0, 0, 0, 0])
            status_col = start_col + 1 + i * 4  # Proper column offset for this health status
            for j, val in enumerate(values):
                cell = ws.cell(row=row, column=status_col + j)
                cell.value = val

                if j == 1:  # Percent column
                    cell.number_format = "0.00%"
        row += 1
    return row

def write_distribution_table(ws, gene_data, health_statuses, start_col):
    start_row = 1
    ws.cell(row=start_row, column=start_col).value = "Percent samples"
    ws.cell(row=start_row, column=start_col).font = Font(bold=True)

    for i, status in enumerate(health_statuses):
        ws.cell(row=start_row, column=start_col + 1 + i).value = status
        ws.cell(row=start_row, column=start_col + 1 + i).font = Font(bold=True)

    status_gene_percents = {status: [] for status in health_statuses}
    for gene in gene_data:
        for status in health_statuses:
            percent = gene_data[gene].get(status, [0, 0])[1]
            if isinstance(percent, (int, float)):
                status_gene_percents[status].append(percent)

    bin_labels = list(bin_gene_percents([]).keys())
    for i, label in enumerate(bin_labels, start=start_row + 1):
        ws.cell(row=i, column=start_col).value = label
        ws.cell(row=i, column=start_col).font = Font(bold=True)

        for j, status in enumerate(health_statuses):
            bins = bin_gene_percents(status_gene_percents[status])
            ws.cell(row=i, column=start_col + 1 + j).value = bins[label]

def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

def main():
    file_path = select_excel_file()
    if not file_path:
        print("No file selected.")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    summary_ws = wb.create_sheet("Summary")

    gene_data, count_labels, health_statuses = extract_gene_summary_data(wb)

    last_summary_row = write_summary_table(summary_ws, gene_data, count_labels, health_statuses, start_col=1)

    summary_col_width = 1 + 4 * len(health_statuses)
    write_distribution_table(summary_ws, gene_data, health_statuses, start_col=summary_col_width + 2)

    autosize_columns(summary_ws)

    wb.save(file_path)
    print(f"Summary and distribution table added to: {file_path}")

if __name__ == "__main__":
    main()
