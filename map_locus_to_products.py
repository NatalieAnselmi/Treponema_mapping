import pandas as pd
from tkinter import Tk, filedialog
from Bio import SeqIO
from openpyxl import load_workbook

###
#Opens an excel file containing top50 and top100 gene lists
#Opens an gbff annotated genome file
#Creates two new sheets in the workbook showing each gene locus next to product
###

# --- Step 1: Select files ---
root = Tk()
root.withdraw()

# Select Excel workbook
top_genes_file = filedialog.askopenfilename(
    title="Select Top Genes Excel Workbook",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
if not top_genes_file:
    raise SystemExit("No Top Genes file selected.")

# Select .gbff genome file
gbff_file = filedialog.askopenfilename(
    title="Select GenBank File (.gbff)",
    filetypes=[("GenBank files", "*.gbff"), ("All files", "*.*")]
)
if not gbff_file:
    raise SystemExit("No GenBank file selected.")

# --- Step 2: Parse gbff and build locus_tag → product dictionary ---
locus_to_product = {}
for record in SeqIO.parse(gbff_file, "genbank"):
    for feature in record.features:
        if feature.type == "CDS":
            locus_tags = feature.qualifiers.get("locus_tag", [])
            products = feature.qualifiers.get("product", [])
            if locus_tags and products:
                locus_to_product[locus_tags[0]] = products[0]

print(f"✅ Parsed {len(locus_to_product)} locus→product mappings from {gbff_file}")

# --- Step 3: Process Top50s and Top100s ---
target_sheets = ["Top50s", "Top100s"]

# Load workbook so we can check/remove old sheets
book = load_workbook(top_genes_file)

with pd.ExcelWriter(top_genes_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    for sheet in target_sheets:
        df = pd.read_excel(top_genes_file, sheet_name=sheet)

        combined = pd.DataFrame()
        for col in df.columns:
            # Keep locus IDs
            combined[f"{col} (Loci)"] = df[col]

            # Map to product name
            combined[f"{col} (Product)"] = df[col].map(
                lambda x: locus_to_product.get(x, "No product found") if pd.notna(x) else x
            )

        # Remove old sheet if it exists
        if f"{sheet}_products" in book.sheetnames:
            std = book[f"{sheet}_products"]
            book.remove(std)

        # Write new sheet
        combined.to_excel(writer, sheet_name=f"{sheet}_products", index=False)

print(f"✅ Added Top50s_products and Top100s_products sheets to {top_genes_file}")
